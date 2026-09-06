import csv
import time

import export_leads as svc


def test_export_leads_writes_csv_with_expected_columns(tmp_path, monkeypatch):
    exports_dir = tmp_path / "exports"
    monkeypatch.setattr(svc, "EXPORTS_DIR", exports_dir)
    monkeypatch.setattr(
        svc.combined_leads, "get_combined_leads",
        lambda filter=None: [{
            "firma": "Muster GmbH", "kontakt": "Max Muster <max@muster.de>", "quelle": "beide",
            "status": "qualifiziert", "score": "8", "letzter_kontakt": "2026-09-01",
            "close_lead_id": "lead_1", "close_link": "https://app.close.com/lead/lead_1/",
            "vault_path": "2026-09-01-Muster-GmbH.md",
        }],
    )

    result = svc.export_leads({}, "csv")

    assert result["ok"] is True
    assert result["anzahl_leads"] == 1
    assert result["download_url"] == f"/api/lead-agent/exports/{result['filename']}"
    written = exports_dir / result["filename"]
    assert written.exists()

    with written.open(encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    assert rows[0] == ["Firma", "Kontakt", "Quelle", "Status", "Score", "Letzter Kontakt", "Close-Lead-ID", "Close-Link", "Vault-Pfad"]
    assert rows[1][0] == "Muster GmbH"


def test_export_leads_writes_real_xlsx_file(tmp_path, monkeypatch):
    import openpyxl

    exports_dir = tmp_path / "exports"
    monkeypatch.setattr(svc, "EXPORTS_DIR", exports_dir)
    monkeypatch.setattr(
        svc.combined_leads, "get_combined_leads",
        lambda filter=None: [{"firma": "Excel Firma", "kontakt": "", "quelle": "vault", "status": "neu",
                               "score": "", "letzter_kontakt": "", "close_lead_id": "", "close_link": "", "vault_path": "x.md"}],
    )

    result = svc.export_leads({}, "xlsx")

    written = exports_dir / result["filename"]
    assert written.exists()
    wb = openpyxl.load_workbook(written)
    ws = wb.active
    assert ws["A1"].value == "Firma"
    assert ws["A1"].font.bold is True
    assert ws["A2"].value == "Excel Firma"


def test_export_leads_rejects_unknown_format(tmp_path, monkeypatch):
    monkeypatch.setattr(svc, "EXPORTS_DIR", tmp_path / "exports")

    result = svc.export_leads({}, "pdf")

    assert result["ok"] is False
    assert "pdf" in result["error"]


def test_export_leads_cleans_up_files_older_than_24h(tmp_path, monkeypatch):
    exports_dir = tmp_path / "exports"
    exports_dir.mkdir()
    old_file = exports_dir / "leads_export_old.csv"
    old_file.write_text("stale", encoding="utf-8")
    old_time = time.time() - (25 * 60 * 60)
    import os
    os.utime(old_file, (old_time, old_time))

    monkeypatch.setattr(svc, "EXPORTS_DIR", exports_dir)
    monkeypatch.setattr(svc.combined_leads, "get_combined_leads", lambda filter=None: [])

    svc.export_leads({}, "csv")

    assert not old_file.exists()


def test_export_leads_keeps_recent_files_during_cleanup(tmp_path, monkeypatch):
    exports_dir = tmp_path / "exports"
    exports_dir.mkdir()
    recent_file = exports_dir / "leads_export_recent.csv"
    recent_file.write_text("fresh", encoding="utf-8")

    monkeypatch.setattr(svc, "EXPORTS_DIR", exports_dir)
    monkeypatch.setattr(svc.combined_leads, "get_combined_leads", lambda filter=None: [])

    svc.export_leads({}, "csv")

    assert recent_file.exists()
