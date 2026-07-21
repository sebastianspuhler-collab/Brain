#!/usr/bin/env python3
"""EINE Excel-Datei mit 12 Monatsblaettern. NUR Nettowerte (kein Brutto).
Einlagen/Entnahmen der Gesellschafter sind KEIN Gewinn einer GbR und werden
komplett ausgeschlossen - sie erscheinen nirgends in diesen Tabellen."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "/Users/sesp01-user/vault/Prozessia-Brain/Finanzen/Belegauswertung_2025/out"
OUTFILE = f"{BASE}/ergebnis_2025_monatlich_netto.xlsx"

merged = json.load(open(f"{BASE}/04_merged.json", encoding='utf-8'))
tx_all = merged['transaktionen']
belege_by_id = {b['id']: b for b in merged['belege']}

RICHTUNG_LABEL = {"AUSGANG": "UMSATZ (Geld rein)", "EINGANG": "AUSGABE (Geld raus)"}
MONATSNAMEN = ["", "01_Januar", "02_Februar", "03_Maerz", "04_April", "05_Mai", "06_Juni",
               "07_Juli", "08_August", "09_September", "10_Oktober", "11_November", "12_Dezember"]

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

def style_and_fit(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = "A2"
    for c in range(1, ncols + 1):
        col = get_column_letter(c)
        maxlen = 10
        for cell in ws[col]:
            if cell.value is not None:
                maxlen = max(maxlen, min(60, len(str(cell.value)) + 2))
        ws.column_dimensions[col].width = maxlen

# NUR geschaeftliche Zahlungen - Einlagen/Entnahmen (Gesellschafter-Privatkonto) und
# Umbuchungen sind kein Umsatz/Gewinn einer GbR und werden hier NICHT aufgefuehrt.
geschaeftlich = [t for t in tx_all if t['kategorie'] == 'GESCHAEFTLICH']

headers = ["Zahlungsdatum", "Richtung", "Betrag netto", "Gegenpartei", "Verwendungszweck",
           "Beleg-Datei", "Status", "Hinweis"]

wb = Workbook()
wb.remove(wb.active)

jahres_umsatz_netto = 0.0
jahres_ausgaben_netto = 0.0
jahres_anzahl_ohne_netto = 0

for m in range(1, 13):
    ws = wb.create_sheet(MONATSNAMEN[m])
    ws.append(headers)

    month_tx = [t for t in geschaeftlich if t['monat'] == m]
    sum_netto_umsatz = sum_netto_ausgabe = 0.0
    n_ohne_netto = 0

    for t in sorted(month_tx, key=lambda x: x['zahlungsdatum']):
        beleg = belege_by_id.get(t['beleg_ids'][0]) if t.get('beleg_ids') else None
        netto = beleg.get('betrag_netto') if beleg else None

        if netto is None:
            n_ohne_netto += 1
            hinweis = "KEIN NETTO VERFUEGBAR - " + (
                "kein Beleg gefunden" if not beleg else "Beleg ohne Netto-/USt-Angabe"
            )
        else:
            hinweis = None

        ws.append([t['zahlungsdatum'], RICHTUNG_LABEL[t['richtung']], netto, t['gegenpartei'],
                   t['verwendungszweck'], beleg['quellref'].split('/')[-1] if beleg else None,
                   t['status'], t.get('pruefgrund') or hinweis])

        if netto is not None:
            if t['richtung'] == 'AUSGANG':
                sum_netto_umsatz += netto
            else:
                sum_netto_ausgabe += netto

    r = ws.max_row + 2
    rows_summary = [
        ("SUMME Umsatz netto:", round(sum_netto_umsatz, 2)),
        ("SUMME Ausgaben netto:", round(sum_netto_ausgabe, 2)),
        ("GEWINN NETTO (Monat):", round(sum_netto_umsatz - sum_netto_ausgabe, 2)),
        ("Anzahl Zahlungen OHNE Netto-Wert (nicht in Summe enthalten):", n_ohne_netto),
    ]
    for label, val in rows_summary:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=3, value=val).font = Font(bold=True)
        r += 1

    style_and_fit(ws, len(headers))
    jahres_umsatz_netto += sum_netto_umsatz
    jahres_ausgaben_netto += sum_netto_ausgabe
    jahres_anzahl_ohne_netto += n_ohne_netto
    print(f"{MONATSNAMEN[m]}: {len(month_tx)} geschaeftliche Zahlungen, {n_ohne_netto} ohne Netto, "
          f"Gewinn netto {round(sum_netto_umsatz - sum_netto_ausgabe, 2)} EUR")

wb.save(OUTFILE)
print()
print("Datei gespeichert:", OUTFILE)
print(f"Jahressumme Umsatz netto: {round(jahres_umsatz_netto,2)} EUR")
print(f"Jahressumme Ausgaben netto: {round(jahres_ausgaben_netto,2)} EUR")
print(f"Jahres-Gewinn netto: {round(jahres_umsatz_netto - jahres_ausgaben_netto,2)} EUR")
print(f"Zahlungen insgesamt ohne Netto-Wert (fehlen in der Summe): {jahres_anzahl_ohne_netto}")
print("Einlagen/Entnahmen der Gesellschafter sind NICHT enthalten (separates Tab 'Einlagen_Entnahmen' in ergebnis_2025.xlsx).")
