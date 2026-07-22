#!/usr/bin/env python3
"""EINE Excel-Datei mit 12 Monatsblaettern. Nur Nettowerte, einfache Spalten,
Gewinn/Verlust pro Monat. Einlagen/Entnahmen der Gesellschafter sind KEIN Gewinn
einer GbR und werden komplett ausgeschlossen - sie erscheinen nirgends hier."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "/Users/sesp01-user/vault/Prozessia-Brain/Finanzen/Belegauswertung_2024/out"
OUTFILE = f"{BASE}/ergebnis_2024_monatlich_netto.xlsx"

merged = json.load(open(f"{BASE}/04_merged.json", encoding='utf-8'))
tx_all = merged['transaktionen']
belege_by_id = {b['id']: b for b in merged['belege']}

RICHTUNG_LABEL = {"AUSGANG": "Umsatz", "EINGANG": "Ausgabe"}
MONATSNAMEN = ["", "01_Januar", "02_Februar", "03_Maerz", "04_April", "05_Mai", "06_Juni",
               "07_Juli", "08_August", "09_September", "10_Oktober", "11_November", "12_Dezember"]

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MISSING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MISSING_FONT = Font(color="9C0006")
EXCLUDED_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
EXCLUDED_FONT = Font(color="808080", italic=True)
USTSATZ = 0.19  # Annahme fuer Zahlungen ohne Beleg, um sie trotzdem in der Nettorechnung zu erfassen

BAGATELLE_AUSSCHLUSS_GRENZE = 10.0
DURCHLAUFPOSTEN_TX_IDS = set()  # 2024: keine bekannten Durchlaufposten

def ausschlussgrund(t, beleg):
    if t['tx_id'] in DURCHLAUFPOSTEN_TX_IDS:
        return "Durchlaufposten (Rundlauf-Buchung)"
    if 'finanzamt' in (t['gegenpartei'] or '').lower():
        return "Finanzamt (steuerneutral)"
    if t['betrag_brutto'] < BAGATELLE_AUSSCHLUSS_GRENZE:
        if not beleg or beleg.get('betrag_netto') is None:
            return f"Bagatelle < {BAGATELLE_AUSSCHLUSS_GRENZE:.0f} EUR ohne Beleg"
    return None

def style_and_fit(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = "A2"
    for c in range(1, ncols + 1):
        col = get_column_letter(c)
        maxlen = 10
        for cell in ws[col]:
            if cell.value is not None:
                maxlen = max(maxlen, min(60, len(str(cell.value)) + 2))
        ws.column_dimensions[col].width = maxlen

# NUR geschaeftliche Zahlungen. Einlagen/Entnahmen (Gesellschafter-Privatkonto) und
# Umbuchungen sind kein Umsatz/Gewinn einer GbR und tauchen hier gar nicht auf.
geschaeftlich = [t for t in tx_all if t['kategorie'] == 'GESCHAEFTLICH']

headers = ["Datum", "Richtung", "Netto-Betrag", "Zaehlt zur Wertung?", "Beleg da?", "Beleg-Dateiname", "Partner", "Verwendungszweck"]

wb = Workbook()
wb.remove(wb.active)

jahr_umsatz = jahr_ausgabe = 0.0
jahr_anzahl_kein_beleg = 0

for m in range(1, 13):
    ws = wb.create_sheet(MONATSNAMEN[m])
    ws.append(headers)

    month_tx = [t for t in geschaeftlich if t['monat'] == m]
    sum_umsatz = sum_ausgabe = 0.0
    n_kein_beleg = 0

    for t in sorted(month_tx, key=lambda x: x['zahlungsdatum']):
        beleg = belege_by_id.get(t['beleg_ids'][0]) if t.get('beleg_ids') else None
        beleg_datei = beleg['quellref'].split('/')[-1] if beleg else ""
        fehlt = beleg is None
        ausschluss = ausschlussgrund(t, beleg)
        if fehlt:
            n_kein_beleg += 1
            netto = round(t['betrag_brutto'] / (1 + USTSATZ), 2)
            beleg_da = "NEIN - fehlt"
        else:
            netto = beleg.get('betrag_netto')
            beleg_da = "Ja"
            if beleg.get('netto_ist_einzelwert'):
                beleg_da = "Ja (nur 1 Betrag, kein Netto/USt getrennt)"

        wertung_label = f"Nein - {ausschluss}" if ausschluss else "Ja"

        row_idx = ws.max_row + 1
        ws.append([t['zahlungsdatum'], RICHTUNG_LABEL[t['richtung']], netto, wertung_label, beleg_da,
                   beleg_datei, t['gegenpartei'], t['verwendungszweck']])
        if ausschluss:
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.fill = EXCLUDED_FILL
                cell.font = EXCLUDED_FONT
        elif fehlt:
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.fill = MISSING_FILL
                cell.font = MISSING_FONT

        if not ausschluss and netto is not None:
            if t['richtung'] == 'AUSGANG':
                sum_umsatz += netto
            else:
                sum_ausgabe += netto

    r = ws.max_row + 2
    for label, val in [
        ("Umsatz netto:", round(sum_umsatz, 2)),
        ("Ausgabe netto:", round(sum_ausgabe, 2)),
        ("GEWINN/VERLUST:", round(sum_umsatz - sum_ausgabe, 2)),
        ("davon Zahlungen ohne Beleg (rot, Netto geschaetzt mit 19% USt):", n_kein_beleg),
    ]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=3, value=val).font = Font(bold=True)
        r += 1

    style_and_fit(ws, len(headers))
    jahr_umsatz += sum_umsatz
    jahr_ausgabe += sum_ausgabe
    jahr_anzahl_kein_beleg += n_kein_beleg
    print(f"{MONATSNAMEN[m]}: Gewinn/Verlust {round(sum_umsatz - sum_ausgabe, 2)} EUR "
          f"({n_kein_beleg} Zahlungen ohne Beleg)")

wb.save(OUTFILE)
print()
print("Datei:", OUTFILE)
print(f"Jahr Umsatz netto: {round(jahr_umsatz,2)} EUR | Ausgabe netto: {round(jahr_ausgabe,2)} EUR "
      f"| GEWINN: {round(jahr_umsatz - jahr_ausgabe,2)} EUR")
print(f"Zahlungen ohne Beleg insgesamt: {jahr_anzahl_kein_beleg}")
