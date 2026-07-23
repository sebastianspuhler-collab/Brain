#!/usr/bin/env python3
"""Schritt 6: Excel-Ausgabe (10 Tabs) + PRUEFFAELLE.md + Konsolenausgabe."""
import json, re
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "/Users/sesp01-user/vault/Prozessia-Brain/Finanzen/Belegauswertung_2024/out"

merged = json.load(open(f"{BASE}/04_merged.json", encoding='utf-8'))
tx_all = merged['transaktionen']
belege_all = merged['belege']
belege_by_id = {b['id']: b for b in belege_all}
auswertung = json.load(open(f"{BASE}/05_auswertung.json", encoding='utf-8'))
ausserhalb = json.load(open(f"{BASE}/ausserhalb_zeitraum.json", encoding='utf-8'))
drive_info = json.load(open(f"{BASE}/02b_belege_drive.json", encoding='utf-8'))

BAGATELLE = 15.0
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

RICHTUNG_LABEL = {
    "AUSGANG": "UMSATZ (Geld rein)",
    "EINGANG": "AUSGABE (Geld raus)",
}
MONATSNAMEN = ["", "Januar", "Februar", "Maerz", "April", "Mai", "Juni", "Juli",
               "August", "September", "Oktober", "November", "Dezember"]

def normalize_partner(name):
    if not name:
        return ""
    n = name.lower()
    n = re.sub(r'\(haftungsbeschr[aä]nkt\)', '', n)
    n = re.sub(r'\b(gmbh|ag|ltd|inc|gbr|ug|egbr|se|bv|b\.v\.|llc|kg|co)\b', '', n)
    n = re.sub(r'[^a-z0-9äöüß]+', ' ', n)
    return re.sub(r'\s+', ' ', n).strip()

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = "A2"

def autofit(ws, ncols, minw=10, maxw=60):
    for c in range(1, ncols + 1):
        col = get_column_letter(c)
        maxlen = minw
        for cell in ws[col]:
            if cell.value is not None:
                maxlen = max(maxlen, min(maxw, len(str(cell.value)) + 2))
        ws.column_dimensions[col].width = maxlen

wb = Workbook()

# ---------- Tab 1: Jahresuebersicht ----------
ws = wb.active
ws.title = "Jahresuebersicht"
jahr = auswertung['jahr']
rows = [
    ("HAUPTKENNZAHLEN (Netto-Basis, nur aus Belegen mit expliziter Netto-Angabe)", ""),
    ("Umsatz netto", jahr['umsatz_netto']),
    ("USt vereinnahmt", jahr['ust_vereinnahmt']),
    ("Ausgaben netto", jahr['ausgaben_netto']),
    ("Vorsteuer abziehbar", jahr['vorsteuer_abziehbar']),
    ("USt-Zahllast (vereinnahmt - abziehbar)", jahr['ust_zahllast']),
    ("Gewinn netto (Umsatz netto - Ausgaben netto)", jahr['gewinn_netto']),
    ("Gewerbeertrag Basis (= Gewinn netto, RECHNERISCH)", jahr['gewerbeertrag_basis']),
    ("Gewerbeertrag nach Freibetrag 24.500 EUR (RECHNERISCH)", jahr['gewerbeertrag_nach_freibetrag']),
    ("Gewinnanteil je Gesellschafter, netto (50/50)", jahr['gewinnanteil_je_gesellschafter']),
    ("", ""),
    ("ZUR EINORDNUNG (Brutto, vollstaendig, aber ohne Netto-/USt-Aufschluesselung)", ""),
    ("Umsatz brutto (alle Zahlungen)", jahr['umsatz_brutto']),
    ("Ausgaben brutto (alle Zahlungen)", jahr['ausgaben_brutto']),
    ("Brutto ohne Netto-/USt-Aufteilung (kein/unvollstaendiger Beleg - fehlt oben im Netto-Gewinn)", jahr['brutto_ohne_aufteilung']),
    ("Vorsteuer gefaehrdet (geschaetzt, mangels Beleg nicht abziehbar)", jahr['vorsteuer_gefaehrdet']),
    ("Anzahl Pruefaelle (Transaktionsebene)", jahr['anzahl_pruefaelle']),
]
for r in rows:
    ws.append(r)
ws.append(("", ""))
ws.append(("HINWEIS:", jahr['hinweis']))
ws.append(("HINWEIS:", "Schritt 3 (Gmail) ist ausgefallen: 'Request had insufficient authentication scopes' - OAuth-Berechtigung fuer Gmail fehlt. Keine E-Mail-Belege in dieser Auswertung enthalten."))
ws.append(("HINWEIS:", "Schritt 1 (Lexoffice) hat keine Bank-Transaktions-API - Transaktionsgrundlage ist der vom Nutzer bereitgestellte Finom-Kontoauszug (Finom_statement_21072026.pdf, 29.10.2024-21.07.2026), auf 2024 gefiltert."))
ws.append(("HINWEIS:", "USt-Voranmeldungen 2024 werden vom Nutzer selbst geprueft, kein Abgleich in diesem Lauf."))
ws.append(("HINWEIS:", "Diese Auswertung ist eine Datengrundlage, KEINE Steuerberatung."))
style_header(ws, 2)
autofit(ws, 2, maxw=90)

# ---------- Tab 2: Monatsuebersicht ----------
ws = wb.create_sheet("Monatsuebersicht")
headers = ["Monat", "Umsatz netto", "Umsatz brutto", "Ausgaben netto", "Ausgaben brutto",
           "Gewinn netto", "USt vereinnahmt", "Vorsteuer abziehbar", "Vorsteuer gefaehrdet",
           "USt-Zahllast", "Brutto ohne Aufteilung", "Anzahl Pruefaelle"]
ws.append(headers)
monate = auswertung['monate']
for m in range(1, 13):
    row = monate[str(m)]
    ws.append([m, row['umsatz_netto'], row['umsatz_brutto'], row['ausgaben_netto'], row['ausgaben_brutto'],
               row['gewinn_netto'], row['ust_vereinnahmt'], row['vorsteuer_abziehbar'],
               row['vorsteuer_gefaehrdet'], row['ust_zahllast'], row['brutto_ohne_aufteilung'],
               row['anzahl_pruefaelle']])
ws.append(["Jahr"] + [jahr[k] for k in ['umsatz_netto','umsatz_brutto','ausgaben_netto','ausgaben_brutto',
           'gewinn_netto','ust_vereinnahmt','vorsteuer_abziehbar','vorsteuer_gefaehrdet',
           'ust_zahllast','brutto_ohne_aufteilung','anzahl_pruefaelle']])
for c in range(1, len(headers) + 1):
    ws.cell(row=14, column=c).font = Font(bold=True)
style_header(ws, len(headers))
autofit(ws, len(headers))

# ---------- Tab 3: Transaktionen ----------
ws = wb.create_sheet("Transaktionen")
headers = ["tx_id", "Zahlungsdatum", "Monat", "Richtung", "Betrag brutto", "Gegenpartei",
           "Verwendungszweck", "Kategorie", "Fremdwaehrung", "Beleg-Dateien", "Status", "Pruefgrund"]
ws.append(headers)
for t in sorted(tx_all, key=lambda x: x['zahlungsdatum']):
    beleg_files = "; ".join(belege_by_id[bid]['quellref'].split('/')[-1] for bid in t.get('beleg_ids', []) if bid in belege_by_id)
    ws.append([t['tx_id'], t['zahlungsdatum'], t['monat'], RICHTUNG_LABEL[t['richtung']], t['betrag_brutto'],
               t['gegenpartei'], t['verwendungszweck'], t['kategorie'], t.get('fremdwaehrung', False),
               beleg_files, t['status'], t.get('pruefgrund')])
style_header(ws, len(headers))
autofit(ws, len(headers))

# ---------- Tabs: ein Blatt pro Monat mit ALLEN Transaktionen (Wunsch Nutzer) ----------
for m in range(1, 13):
    sheet_name = f"{m:02d}_{MONATSNAMEN[m]}"
    ws = wb.create_sheet(sheet_name)
    headers = ["Zahlungsdatum", "Richtung", "Betrag brutto", "Gegenpartei", "Verwendungszweck",
               "Kategorie", "Beleg-Datei(en)", "Status", "Pruefgrund"]
    ws.append(headers)
    month_tx = [t for t in tx_all if t['monat'] == m]
    for t in sorted(month_tx, key=lambda x: x['zahlungsdatum']):
        beleg_files = "; ".join(belege_by_id[bid]['quellref'].split('/')[-1] for bid in t.get('beleg_ids', []) if bid in belege_by_id)
        ws.append([t['zahlungsdatum'], RICHTUNG_LABEL[t['richtung']], t['betrag_brutto'], t['gegenpartei'],
                   t['verwendungszweck'], t['kategorie'], beleg_files, t['status'], t.get('pruefgrund')])
    # Summenzeile
    umsatz = sum(t['betrag_brutto'] for t in month_tx if t['richtung'] == 'AUSGANG' and t['kategorie'] == 'GESCHAEFTLICH')
    ausgabe = sum(t['betrag_brutto'] for t in month_tx if t['richtung'] == 'EINGANG' and t['kategorie'] == 'GESCHAEFTLICH')
    einlage = sum(t['betrag_brutto'] for t in month_tx if t['richtung'] == 'AUSGANG' and t['kategorie'] == 'EINLAGE_ENTNAHME')
    entnahme = sum(t['betrag_brutto'] for t in month_tx if t['richtung'] == 'EINGANG' and t['kategorie'] == 'EINLAGE_ENTNAHME')
    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="SUMME Umsatz (geschaeftlich, Geld rein):").font = Font(bold=True)
    ws.cell(row=r, column=3, value=round(umsatz, 2)).font = Font(bold=True)
    ws.cell(row=r + 1, column=1, value="SUMME Ausgaben (geschaeftlich, Geld raus):").font = Font(bold=True)
    ws.cell(row=r + 1, column=3, value=round(ausgabe, 2)).font = Font(bold=True)
    ws.cell(row=r + 2, column=1, value="SUMME Einlagen (Gesellschafter, Geld rein):").font = Font(bold=True)
    ws.cell(row=r + 2, column=3, value=round(einlage, 2)).font = Font(bold=True)
    ws.cell(row=r + 3, column=1, value="SUMME Entnahmen (Gesellschafter, Geld raus):").font = Font(bold=True)
    ws.cell(row=r + 3, column=3, value=round(entnahme, 2)).font = Font(bold=True)
    style_header(ws, len(headers))
    autofit(ws, len(headers))

# ---------- Tab: Umsaetze_2024 (chronologisch, mit Beleg-Zuordnungsversuch) ----------
ws = wb.create_sheet("Umsaetze_2024")
headers = ["Zahlungsdatum", "Betrag brutto", "Gegenpartei", "Verwendungszweck",
           "Beleg gefunden?", "Beleg-Datei", "Rechnungsnummer (Beleg)", "Status", "Pruefgrund/Hinweis"]
ws.append(headers)
umsaetze = [t for t in tx_all if t['kategorie'] == 'GESCHAEFTLICH' and t['richtung'] == 'AUSGANG']
for t in sorted(umsaetze, key=lambda x: x['zahlungsdatum']):
    if t.get('beleg_ids'):
        b = belege_by_id.get(t['beleg_ids'][0])
        beleg_gefunden = "JA"
        beleg_datei = b['quellref'].split('/')[-1] if b else None
        beleg_nr = b.get('rechnungsnummer') if b else None
    else:
        beleg_gefunden = "NEIN"
        beleg_datei = None
        beleg_nr = None
    ws.append([t['zahlungsdatum'], t['betrag_brutto'], t['gegenpartei'], t['verwendungszweck'],
               beleg_gefunden, beleg_datei, beleg_nr, t['status'], t.get('pruefgrund')])
style_header(ws, len(headers))
autofit(ws, len(headers))

# ---------- Tab: Statistik_fehlende_Rechnungen ----------
ws = wb.create_sheet("Statistik_fehlende_Rechnungen")
headers = ["Gegenpartei (normalisiert)", "Anzahl Zahlungen gesamt", "davon ohne Beleg (BELEG_FEHLT)",
           "Anteil ohne Beleg", "Summe ohne Beleg (EUR)", "Richtung"]
ws.append(headers)
geschaeftlich = [t for t in tx_all if t['kategorie'] == 'GESCHAEFTLICH']
stat = defaultdict(lambda: {'gesamt': 0, 'fehlt': 0, 'summe_fehlt': 0.0, 'richtung': set(), 'name': None})
for t in geschaeftlich:
    key = normalize_partner(t['gegenpartei'])
    if not key:
        continue
    s = stat[key]
    s['gesamt'] += 1
    s['richtung'].add(RICHTUNG_LABEL[t['richtung']])
    s['name'] = s['name'] or t['gegenpartei']
    if t['status'] == 'BELEG_FEHLT':
        s['fehlt'] += 1
        s['summe_fehlt'] += t['betrag_brutto']
rows_stat = []
for key, s in stat.items():
    if s['fehlt'] == 0:
        continue
    anteil = round(100 * s['fehlt'] / s['gesamt'], 0)
    rows_stat.append((s['name'], s['gesamt'], s['fehlt'], f"{anteil:.0f}%", round(s['summe_fehlt'], 2),
                       " / ".join(sorted(s['richtung']))))
rows_stat.sort(key=lambda r: -r[4])
for r in rows_stat:
    ws.append(list(r))
style_header(ws, len(headers))
autofit(ws, len(headers))

# ---------- Tab 4: Belege_ohne_Zahlung ----------
ws = wb.create_sheet("Belege_ohne_Zahlung")
headers = ["Datei", "Partner", "Rechnungsdatum", "Rechnungsnummer", "Betrag brutto", "Waehrung",
           "Richtung", "Vermuteter Zahlungsweg", "Pruefgrund"]
ws.append(headers)
ohne_zahlung = [b for b in belege_all if b['status'] == 'OHNE_ZAHLUNG' and (b.get('betrag_brutto') or 0) > BAGATELLE]
for b in sorted(ohne_zahlung, key=lambda x: -(x.get('betrag_brutto') or 0)):
    ws.append([b['quellref'].split('/')[-1], b.get('partner'), b.get('rechnungsdatum'),
               b.get('rechnungsnummer'), b.get('betrag_brutto'), b.get('waehrung'),
               b.get('richtung'), b.get('zahlungsweg'), b.get('pruefgrund')])
style_header(ws, len(headers))
autofit(ws, len(headers))

# ---------- Tab 5: Zahlungen_ohne_Beleg ----------
ws = wb.create_sheet("Zahlungen_ohne_Beleg")
headers = ["Zahlungsdatum", "Betrag brutto", "Richtung", "Gegenpartei", "Verwendungszweck"]
ws.append(headers)
ohne_beleg = [t for t in tx_all if t['status'] == 'BELEG_FEHLT']
for t in sorted(ohne_beleg, key=lambda x: -x['betrag_brutto']):
    ws.append([t['zahlungsdatum'], t['betrag_brutto'], RICHTUNG_LABEL[t['richtung']], t['gegenpartei'], t['verwendungszweck']])
style_header(ws, len(headers))
autofit(ws, len(headers))

# ---------- Tab 6: Prueffaelle (echte, keine Bagatelle) ----------
ws = wb.create_sheet("Prueffaelle")
headers = ["Typ", "Datum", "Betrag brutto", "Partner/Gegenpartei", "Quelle", "Status", "Rueckfrage/Pruefgrund"]
ws.append(headers)
pruef_tx = [t for t in tx_all if t['status'] == 'PRUEFFALL']
pruef_belege = [b for b in belege_all if b['status'] == 'PRUEFFALL']
pruef_rows = []
for t in pruef_tx:
    pruef_rows.append(("Transaktion", t['zahlungsdatum'], t['betrag_brutto'], t['gegenpartei'],
                        f"Finom-Kontoauszug ({t['tx_id']})", t['status'], t.get('pruefgrund')))
for b in pruef_belege:
    pruef_rows.append(("Beleg", b.get('rechnungsdatum'), b.get('betrag_brutto'), b.get('partner'),
                        b['quellref'].split('/')[-1], b['status'], b.get('pruefgrund')))
pruef_rows.sort(key=lambda r: -(r[2] or 0))
for r in pruef_rows:
    ws.append(list(r))
style_header(ws, len(headers))
autofit(ws, len(headers))

# ---------- Tab 7: Bagatellfaelle_akzeptiert ----------
ws = wb.create_sheet("Bagatellfaelle_akzeptiert")
headers = ["Typ", "Datum", "Betrag brutto", "Partner/Gegenpartei", "Quelle", "USt-Satz unterstellt"]
ws.append(headers)
bag_tx = [t for t in tx_all if t['status'] == 'AKZEPTIERT_BAGATELLE']
bag_belege = [b for b in belege_all if b['status'] == 'AKZEPTIERT_BAGATELLE']
bag_rows = []
for t in bag_tx:
    bag_rows.append(("Transaktion", t['zahlungsdatum'], t['betrag_brutto'], t['gegenpartei'], t['tx_id'], "19% (unterstellt)"))
for b in bag_belege:
    bag_rows.append(("Beleg", b.get('rechnungsdatum'), b.get('betrag_brutto'), b.get('partner'),
                      b['quellref'].split('/')[-1], "19% (unterstellt)" if b.get('ust_satz_unterstellt') else (f"{b.get('ust_satz')}%" if b.get('ust_satz') is not None else "-")))
bag_rows.sort(key=lambda r: -(r[2] or 0))
for r in bag_rows:
    ws.append(list(r))
style_header(ws, len(headers))
autofit(ws, len(headers))

# ---------- Tab 8: Abos ----------
ws = wb.create_sheet("Abos")
headers = ["Anbieter", "Erkannter Betrag (ca.)", "Intervall", "Erkannte Monate", "Fehlende Monate", "Jahressumme (erkannt)"]
ws.append(headers)
geschaeftlich_eingang = [t for t in tx_all if t['kategorie'] == 'GESCHAEFTLICH' and t['richtung'] == 'EINGANG']
by_partner = defaultdict(list)
for t in geschaeftlich_eingang:
    by_partner[normalize_partner(t['gegenpartei'])].append(t)
for key, txs in sorted(by_partner.items()):
    if not key:
        continue
    # gleicher Betrag (+-0.02) in >=2 Monaten
    by_amount_bucket = defaultdict(set)
    for t in txs:
        bucket = round(t['betrag_brutto'] / 0.02)
        by_amount_bucket[bucket].add(t['monat'])
    for bucket, monate_set in by_amount_bucket.items():
        if len(monate_set) >= 2:
            betrag = bucket * 0.02
            monate_sorted = sorted(monate_set)
            fehlende = [m for m in range(min(monate_sorted), max(monate_sorted) + 1) if m not in monate_set]
            intervall = "monatlich" if len(monate_sorted) >= 3 else "unregelmaessig/monatlich (wenige Datenpunkte)"
            jahressumme = sum(t['betrag_brutto'] for t in txs if round(t['betrag_brutto']/0.02) == bucket)
            ws.append([txs[0]['gegenpartei'], round(betrag, 2), intervall,
                       ", ".join(str(m) for m in monate_sorted),
                       ", ".join(str(m) for m in fehlende) if fehlende else "-",
                       round(jahressumme, 2)])
style_header(ws, len(headers))
autofit(ws, len(headers))

# ---------- Tab 9: Einlagen_Entnahmen ----------
ws = wb.create_sheet("Einlagen_Entnahmen")
headers = ["Datum", "Richtung", "Betrag brutto", "Gesellschafter/Gegenpartei", "Verwendungszweck", "Status"]
ws.append(headers)
ee = [t for t in tx_all if t['kategorie'] == 'EINLAGE_ENTNAHME']
for t in sorted(ee, key=lambda x: x['zahlungsdatum']):
    ws.append([t['zahlungsdatum'], RICHTUNG_LABEL[t['richtung']], t['betrag_brutto'], t['gegenpartei'], t['verwendungszweck'], t['status']])
ws.append(["Summe", "", round(sum(t['betrag_brutto'] for t in ee), 2), "", "", ""])
style_header(ws, len(headers))
autofit(ws, len(headers))

# ---------- Tab 10: Quellenabgleich ----------
ws = wb.create_sheet("Quellenabgleich")
headers = ["Quelle", "Anzahl", "Hinweis"]
ws.append(headers)
n_lex_tx = len(tx_all)
n_lokal = len(belege_all) + len(ausserhalb)
n_matched = sum(1 for t in tx_all if t.get('beleg_ids'))
n_drive_confirmed_dupes = len(drive_info.get('bestaetigte_duplikate_beispiele', []))
n_drive_only = len(drive_info.get('drive_only_gefunden', []))
rows = [
    ("Lexoffice (Bank-API)", 0, "Keine Bank-Transaktions-API vorhanden (nur Voucher/Payment-Endpoints) - siehe Jahresuebersicht-Hinweis"),
    ("Finom-Kontoauszug (Ersatz-Rueckgrat)", n_lex_tx, "Transaktionen 2024, gefiltert aus Gesamtauszug 29.10.2024-21.07.2026"),
    ("Lokaler Ordner (Belege)", n_lokal, f"davon {len(ausserhalb)} ausserhalb 2024 verschoben (out/ausserhalb_zeitraum.json)"),
    ("Google Drive (Finanzordner)", "~60+ Dateien gesichtet", f"Stichprobenartig ueber ~20 Ordner geprueft, {n_drive_confirmed_dupes}+ bestaetigte Duplikate lokaler Belege, {n_drive_only} Drive-only Fund (Finom 2025er-Belege, ausserhalb 2024). Kein vollstaendiger rekursiver Scan aller ~50 Unterordner (siehe out/02b_belege_drive.json)."),
    ("Gmail", 0, "AUSGEFALLEN: 'insufficient authentication scopes' - OAuth-Berechtigung fehlt, keine E-Mail-Belege erfasst"),
    ("Transaktionen mit Beleg verknuepft", n_matched, f"von {n_lex_tx} Transaktionen gesamt"),
    ("Belege insgesamt (nach Dedup, in 2024)", len(belege_all), ""),
]
for r in rows:
    ws.append(list(r))
style_header(ws, len(headers))
autofit(ws, len(headers), maxw=100)

wb.save(f"{BASE}/ergebnis_2024.xlsx")
print("Excel gespeichert:", f"{BASE}/ergebnis_2024.xlsx")
# Bewusst KEINE Kopie in den Jahresordner (Nutzerentscheidung 2026-07-22): dort
# soll nur noch ergebnis_2024_monatlich_netto.xlsx liegen (siehe step7).

# ---------- PRUEFFAELLE.md ----------
pruef_rows.sort(key=lambda r: -(r[2] or 0))
lines = ["# Pruefaelle Belegauswertung Prozessia GbR 2024", "",
         f"Insgesamt {len(pruef_rows)} echte Pruefaelle (Bagatellfaelle <= {BAGATELLE} EUR ausgeschlossen, siehe eigene Liste).", ""]
for i, r in enumerate(pruef_rows, 1):
    typ, datum, betrag, partner, quelle, status, pruefgrund = r
    lines.append(f"## {i}. {partner or '(unbekannt)'} - {betrag if betrag is not None else '?'} EUR ({datum or 'Datum unklar'})")
    lines.append(f"- Typ: {typ}")
    lines.append(f"- Quelle: {quelle}")
    lines.append(f"- Status: {status}")
    lines.append(f"- Rueckfrage: {pruefgrund}")
    lines.append("")
with open(f"{BASE}/PRUEFFAELLE.md", 'w', encoding='utf-8') as f:
    f.write("\n".join(lines))
print("PRUEFFAELLE.md gespeichert:", f"{BASE}/PRUEFFAELLE.md")

# ---------- Konsolenausgabe ----------
status_counter = Counter(t['status'] for t in tx_all)
print()
print("=== KONSOLENAUSGABE ===")
print(f"Transaktionen 2024 gesamt: {len(tx_all)}")
print(f"  BELEGT: {status_counter.get('BELEGT', 0)}")
print(f"  BELEG_FEHLT: {status_counter.get('BELEG_FEHLT', 0)}")
print(f"  AKZEPTIERT_BAGATELLE: {status_counter.get('AKZEPTIERT_BAGATELLE', 0)}")
print(f"  PRUEFFALL (echt): {status_counter.get('PRUEFFALL', 0)}")
print(f"Belege ohne Zahlung (> {BAGATELLE} EUR): {len(ohne_zahlung)}")
print()
print("Jahreskennzahlen (RECHNERISCH, siehe Disclaimer):")
for k in ['umsatz_netto', 'umsatz_brutto', 'ausgaben_netto', 'ausgaben_brutto',
          'gewinn_netto', 'ust_zahllast', 'gewerbeertrag_nach_freibetrag',
          'gewinnanteil_je_gesellschafter']:
    print(f"  {k}: {jahr[k]} EUR")
print()
print("5 groesste echte Pruefaelle:")
for r in pruef_rows[:5]:
    print(f"  {r[1]} | {r[2]} EUR | {r[3]} | {r[6]}")
